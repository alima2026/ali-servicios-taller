
from __future__ import annotations

import re
import unicodedata
from datetime import date
from io import BytesIO
from pathlib import Path

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


st.set_page_config(
    page_title="Dashboard Taller Magna",
    layout="wide",
    initial_sidebar_state="expanded",
)

DEFAULT_EXCEL_PATH = Path(__file__).with_name("vehiculos_en_reparacion_magna.xlsx")
PREFERRED_SHEETS = ["SINIESTROS", "PARTICULAR Y GARANTIAS"]
WORKBOOK_CACHE_VERSION = "2026-04-08-04"

CANONICAL_COLUMNS = [
    "FECHA",
    "CANAL",
    "DIAS EN TALLER",
    "COMPANIA",
    "NRO SINIESTRO",
    "PROVEEDOR",
    "CHASIS",
    "KILOMETRAJE",
    "MATRICULA",
    "NOMBRE CLIENTE",
    "TELEFONO",
    "MARCA",
    "MODELO",
    "CODIGO",
    "REPUESTOS SOLICITADO",
    "MONTO PIEZA",
    "MONTO M OBRA",
    "STATUS DEL REPUESTO",
    "STATUS DEL VEHICULO",
    "FECHA ENTREGA PIEZA",
    "VENTA",
    "MOTIVO",
    "COMENTARIOS",
]

COLUMN_ALIASES = {
    "FECHA": "FECHA",
    "CANAL": "CANAL",
    "DIAS EN TALLER": "DIAS EN TALLER",
    "DIAS TALLER": "DIAS EN TALLER",
    "COMPANIA": "COMPANIA",
    "COMPANIAS": "COMPANIA",
    "N SINIESTRO": "NRO SINIESTRO",
    "NRO SINIESTRO": "NRO SINIESTRO",
    "NRO DE SINIESTRO": "NRO SINIESTRO",
    "NUMERO SINIESTRO": "NRO SINIESTRO",
    "PROVEEDOR": "PROVEEDOR",
    "CHASIS": "CHASIS",
    "KILOMETRAJE": "KILOMETRAJE",
    "KILOMETROS": "KILOMETRAJE",
    "KM": "KILOMETRAJE",
    "MATRICULA": "MATRICULA",
    "NOMBRE CLIENTE": "NOMBRE CLIENTE",
    "TELEFONO": "TELEFONO",
    "MARCA": "MARCA",
    "MODELO": "MODELO",
    "CODIGO": "CODIGO",
    "REPUESTOS SOLICITADO": "REPUESTOS SOLICITADO",
    "REPUESTO SOLICITADO": "REPUESTOS SOLICITADO",
    "MONTO PIEZA": "MONTO PIEZA",
    "MONTO M OBRA": "MONTO M OBRA",
    "MONTO M OBRAS": "MONTO M OBRA",
    "STATUS DEL REPUESTO": "STATUS DEL REPUESTO",
    "STATUS REPUESTO": "STATUS DEL REPUESTO",
    "STATUS DEL VEHICULO": "STATUS DEL VEHICULO",
    "STATUS VEHICULO": "STATUS DEL VEHICULO",
    "FECHA ENTREGA PIEZA": "FECHA ENTREGA PIEZA",
    "VENTA": "VENTA",
    "FECHA VENTA": "VENTA",
    "FECHA DE VENTA": "VENTA",
    "MOTIVO": "MOTIVO",
    "COMENTARIOS": "COMENTARIOS",
}

UNASSIGNED_PROVIDER_LABEL = "SIN PROVEEDOR ASIGNADO"
MIXED_PROVIDER_LABEL = "PROVEEDORES MIXTOS"
PIECE_RESULT_ORDER = ["GANADA MAGNA", "PERDIDA", "SIN PROVEEDOR"]
SEMAFORO_ORDER = ["NORMAL", "ATENCION", "DEMORA ALTA", "CRITICA", "SIN DATO"]
WARRANTY_STATUS_ORDER = [
    "GARANTIA VIGENTE",
    "SIN GARANTIA >100.000 KM",
    "SIN GARANTIA >3 ANOS",
    "SIN GARANTIA >100.000 KM Y >3 ANOS",
    "REVISAR KILOMETRAJE",
    "REVISAR FECHA DE VENTA",
    "REVISAR KILOMETRAJE Y FECHA DE VENTA",
    "NO GARANTIA",
]
MOTIVO_PARTICULAR_ORDER = [
    "MUY CARO Y HAY STOCK",
    "NO HAY STOCK",
    "SIN MOTIVO",
    "OTROS",
]
STOCK_WAIT_ORDER = ["0-40 DIAS", "41-55 DIAS - ATENCION", ">55 DIAS - CRITICA", "SIN DATO"]


def inject_css() -> None:
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 1rem;
            padding-bottom: 2rem;
        }
        .title-card {
            background: linear-gradient(135deg, rgba(255,255,255,0.98) 0%, rgba(240,249,255,0.98) 100%);
            border: 1px solid rgba(15,23,42,0.08);
            border-radius: 24px;
            padding: 1.25rem 1.5rem;
            margin-bottom: 1rem;
            box-shadow: 0 12px 28px rgba(15,23,42,0.08);
        }
        .title-card h1 {
            margin: 0;
            font-size: 2.5rem;
            line-height: 1;
            color: #0f172a;
            letter-spacing: -0.03em;
        }
        .title-card h1 span {
            color: #0f766e;
        }
        .title-card p {
            margin: 0.55rem 0 0 0;
            color: #475569;
            font-size: 1rem;
            font-weight: 500;
        }
        .hero-card {
            background: linear-gradient(135deg, #0f172a 0%, #0f766e 100%);
            color: white;
            border-radius: 22px;
            padding: 1rem 1.2rem;
            box-shadow: 0 10px 24px rgba(15,23,42,0.18);
            margin-bottom: 1rem;
        }
        .hero-title {
            font-size: 1.05rem;
            font-weight: 800;
            margin-bottom: 0.25rem;
        }
        .hero-text {
            font-size: 0.95rem;
            opacity: 0.95;
        }
        .metric-card {
            background: rgba(255,255,255,0.99);
            border: 1px solid rgba(15,23,42,0.06);
            border-radius: 18px;
            padding: 0.95rem 1rem;
            box-shadow: 0 4px 12px rgba(15,23,42,0.05);
            min-height: 112px;
            margin-bottom: 0.65rem;
        }
        .metric-label {
            font-size: 0.88rem;
            color: #64748b;
            margin-bottom: 0.25rem;
            font-weight: 700;
        }
        .metric-value {
            font-size: 1.65rem;
            line-height: 1.1;
            font-weight: 900;
            color: #0f172a;
        }
        .metric-help {
            margin-top: 0.35rem;
            color: #475569;
            font-size: 0.84rem;
        }
        .status-good {
            background: #ecfdf5;
            color: #166534;
            border-left: 6px solid #22c55e;
            padding: 0.85rem 1rem;
            border-radius: 14px;
            font-weight: 700;
            margin-bottom: 1rem;
        }
        .status-mid {
            background: #fffbeb;
            color: #92400e;
            border-left: 6px solid #f59e0b;
            padding: 0.85rem 1rem;
            border-radius: 14px;
            font-weight: 700;
            margin-bottom: 1rem;
        }
        .status-alert {
            background: #fef2f2;
            color: #991b1b;
            border-left: 6px solid #ef4444;
            padding: 0.85rem 1rem;
            border-radius: 14px;
            font-weight: 700;
            margin-bottom: 1rem;
        }
        .tab-note {
            color: #475569;
            font-size: 0.92rem;
            margin-bottom: 0.75rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def slug_text(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def first_non_empty(series: pd.Series) -> object:
    for value in series:
        if isinstance(value, pd.Timestamp) and not pd.isna(value):
            return value
        text = normalize_text(value)
        if text:
            return text
    return pd.NA


def count_non_empty(series: pd.Series) -> int:
    return int(sum(bool(normalize_text(value)) for value in series))


def provider_display_label(value: object) -> str:
    text = normalize_text(value).upper()
    return text if text else UNASSIGNED_PROVIDER_LABEL


def vehicle_provider_label(series: pd.Series) -> str:
    provider_values = [normalize_text(value).upper() for value in series if normalize_text(value)]
    unique_values: list[str] = list(dict.fromkeys(provider_values))
    has_unassigned = int(sum(not normalize_text(value) for value in series)) > 0

    if not unique_values:
        return UNASSIGNED_PROVIDER_LABEL
    if len(unique_values) == 1 and not has_unassigned:
        return unique_values[0]
    if len(unique_values) == 1 and has_unassigned:
        return f"{unique_values[0]} + SIN ASIGNAR"
    if has_unassigned:
        return f"{MIXED_PROVIDER_LABEL} + SIN ASIGNAR"
    return MIXED_PROVIDER_LABEL


def unique_join(series: pd.Series) -> str:
    seen: set[str] = set()
    values: list[str] = []
    for value in series:
        text = normalize_text(value)
        if not text:
            continue
        key = slug_text(text)
        if key in seen:
            continue
        seen.add(key)
        values.append(text)
    return " | ".join(values)


def detect_header_row(raw_df: pd.DataFrame, max_rows: int = 15) -> int:
    for idx in range(min(max_rows, len(raw_df))):
        row_tokens = [slug_text(value) for value in raw_df.iloc[idx].tolist()]
        token_set = set(token for token in row_tokens if token)
        has_fecha = "FECHA" in token_set
        has_chasis = "CHASIS" in token_set
        has_repuesto = any(token.startswith("REPUESTOS SOLICITADO") for token in token_set)
        if has_fecha and has_chasis and has_repuesto:
            return idx
    return 0


def standardize_columns(columns: list[object]) -> list[str]:
    normalized: list[str] = []
    for idx, value in enumerate(columns):
        key = slug_text(value)
        normalized.append(COLUMN_ALIASES.get(key, normalize_text(value) or f"COL_{idx}"))
    return normalized


def classify_piece_result(provider: object) -> str:
    key = slug_text(provider)
    if not key:
        return "SIN PROVEEDOR"
    if key == "MAGNA":
        return "GANADA MAGNA"
    return "PERDIDA"


def classify_semaforo(dias: object) -> str:
    if pd.isna(dias):
        return "SIN DATO"
    dias = float(dias)
    if dias <= 30:
        return "NORMAL"
    if dias <= 45:
        return "ATENCION"
    if dias <= 70:
        return "DEMORA ALTA"
    return "CRITICA"


def normalize_motivo_particular(value: object) -> str:
    key = slug_text(value)
    if not key:
        return "SIN MOTIVO"
    if "NO HAY STOCK" in key or "SIN STOCK" in key:
        return "NO HAY STOCK"
    if "CARO" in key:
        return "MUY CARO Y HAY STOCK"
    return "OTROS"


def classify_stock_wait(dias: object) -> str:
    if pd.isna(dias):
        return "SIN DATO"
    dias = float(dias)
    if dias <= 40:
        return "0-40 DIAS"
    if dias <= 55:
        return "41-55 DIAS - ATENCION"
    return ">55 DIAS - CRITICA"


def classify_warranty_status(canal: object, kilometraje: object, venta: object) -> str:
    canal_key = slug_text(canal)
    if canal_key != "GARANTIA":
        return "NO GARANTIA"

    km_ok = None
    if pd.notna(kilometraje):
        km_ok = float(kilometraje) <= 100000

    venta_ok = None
    if isinstance(venta, pd.Timestamp) and not pd.isna(venta):
        age_days = (pd.Timestamp(date.today()) - venta.normalize()).days
        venta_ok = age_days <= int(365.25 * 3)

    if km_ok is False and venta_ok is False:
        return "SIN GARANTIA >100.000 KM Y >3 ANOS"
    if km_ok is False:
        return "SIN GARANTIA >100.000 KM"
    if venta_ok is False:
        return "SIN GARANTIA >3 ANOS"
    if km_ok is None and venta_ok is None:
        return "REVISAR KILOMETRAJE Y FECHA DE VENTA"
    if km_ok is None:
        return "REVISAR KILOMETRAJE"
    if venta_ok is None:
        return "REVISAR FECHA DE VENTA"
    return "GARANTIA VIGENTE"


def warranty_status_slug(value: object) -> str:
    return slug_text(value)


def clean_taller_dataframe(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    df = df.copy()
    for col in CANONICAL_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA

    ordered_cols = CANONICAL_COLUMNS + [col for col in df.columns if col not in CANONICAL_COLUMNS]
    df = df[ordered_cols]

    text_cols = [col for col in df.columns if col not in {"FECHA", "FECHA ENTREGA PIEZA"}]
    for col in text_cols:
        df[col] = df[col].apply(lambda value: normalize_text(value) if isinstance(value, str) else value)

    fillable_cols = [
        "FECHA",
        "CANAL",
        "DIAS EN TALLER",
        "COMPANIA",
        "NRO SINIESTRO",
        "CHASIS",
        "KILOMETRAJE",
        "MATRICULA",
        "NOMBRE CLIENTE",
        "TELEFONO",
        "MARCA",
        "MODELO",
        "STATUS DEL REPUESTO",
        "STATUS DEL VEHICULO",
        "FECHA ENTREGA PIEZA",
        "VENTA",
        "COMENTARIOS",
    ]
    for col in fillable_cols + ["CODIGO", "REPUESTOS SOLICITADO", "MONTO PIEZA", "MONTO M OBRA"]:
        df[col] = df[col].apply(lambda value: pd.NA if normalize_text(value) == "" else value)

    df["CHASIS"] = df["CHASIS"].ffill()
    df["MATRICULA"] = df["MATRICULA"].ffill()
    df["VEHICULO_ID"] = df["CHASIS"].where(df["CHASIS"].notna(), df["MATRICULA"]).ffill()
    df = df[df["VEHICULO_ID"].notna()].copy()

    grouped = df.groupby("VEHICULO_ID", dropna=False, sort=False)
    for col in [
        "FECHA",
        "CANAL",
        "DIAS EN TALLER",
        "COMPANIA",
        "NRO SINIESTRO",
        "CHASIS",
        "KILOMETRAJE",
        "MATRICULA",
        "NOMBRE CLIENTE",
        "TELEFONO",
        "MARCA",
        "MODELO",
        "STATUS DEL VEHICULO",
        "VENTA",
    ]:
        df[col] = grouped[col].transform(lambda series: series.ffill().bfill())

    df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")
    df["FECHA ENTREGA PIEZA"] = pd.to_datetime(df["FECHA ENTREGA PIEZA"], errors="coerce")
    df["VENTA"] = pd.to_datetime(df["VENTA"], errors="coerce")
    df["DIAS EN TALLER"] = pd.to_numeric(df["DIAS EN TALLER"], errors="coerce")
    df["KILOMETRAJE"] = pd.to_numeric(df["KILOMETRAJE"], errors="coerce")
    df["MONTO PIEZA"] = pd.to_numeric(df["MONTO PIEZA"], errors="coerce")
    df["MONTO M OBRA"] = pd.to_numeric(df["MONTO M OBRA"], errors="coerce")

    for col in ["CANAL", "COMPANIA", "PROVEEDOR", "MARCA", "STATUS DEL REPUESTO", "STATUS DEL VEHICULO"]:
        df[col] = df[col].apply(normalize_text).str.upper()

    for col in [
        "MODELO",
        "CHASIS",
        "MATRICULA",
        "NOMBRE CLIENTE",
        "TELEFONO",
        "CODIGO",
        "REPUESTOS SOLICITADO",
        "COMENTARIOS",
        "NRO SINIESTRO",
        "MOTIVO",
    ]:
        df[col] = df[col].apply(normalize_text)

    fallback_days = (pd.Timestamp(date.today()) - df["FECHA"]).dt.days
    fallback_days = fallback_days.where(fallback_days >= 0)
    df["DIAS EFECTIVOS PIEZA"] = df["DIAS EN TALLER"].where(df["DIAS EN TALLER"].notna(), fallback_days)
    df["MOTIVO_NORMALIZADO"] = df["MOTIVO"].apply(normalize_motivo_particular)
    df["PROVEEDOR_NORMALIZADO"] = df["PROVEEDOR"].apply(slug_text)
    df["PROVEEDOR_DISPLAY"] = df["PROVEEDOR"].apply(provider_display_label)
    df["MAGNA_ADJUDICADO"] = df["PROVEEDOR_NORMALIZADO"] == "MAGNA"
    df["PIEZA_RESULTADO"] = df["PROVEEDOR"].apply(classify_piece_result)
    df["PIEZA_ENTREGADA"] = df["STATUS DEL REPUESTO"].apply(slug_text) == "ENTREGADO"
    df["GARANTIA_ESTADO_PIEZA"] = df.apply(
        lambda row: classify_warranty_status(row.get("CANAL"), row.get("KILOMETRAJE"), row.get("VENTA")),
        axis=1,
    )
    df["SEMAFORO_STOCK_PIEZA"] = df.apply(
        lambda row: classify_stock_wait(row.get("DIAS EFECTIVOS PIEZA"))
        if row.get("MOTIVO_NORMALIZADO") == "NO HAY STOCK"
        else "SIN DATO",
        axis=1,
    )
    df["SOURCE_SHEET"] = sheet_name
    return df.reset_index(drop=True)


def ensure_analysis_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    for col in CANONICAL_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA

    for col in ["PROVEEDOR", "STATUS DEL REPUESTO", "STATUS DEL VEHICULO", "CHASIS", "MATRICULA", "MOTIVO"]:
        if col not in df.columns:
            df[col] = pd.NA

    for col in ["PROVEEDOR", "STATUS DEL REPUESTO", "STATUS DEL VEHICULO", "CHASIS", "MATRICULA", "MOTIVO"]:
        df[col] = df[col].apply(normalize_text)

    df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")
    df["VENTA"] = pd.to_datetime(df["VENTA"], errors="coerce")
    df["FECHA ENTREGA PIEZA"] = pd.to_datetime(df["FECHA ENTREGA PIEZA"], errors="coerce")
    df["DIAS EN TALLER"] = pd.to_numeric(df["DIAS EN TALLER"], errors="coerce")
    df["KILOMETRAJE"] = pd.to_numeric(df["KILOMETRAJE"], errors="coerce")

    if "VEHICULO_ID" not in df.columns:
        chasis = df["CHASIS"].replace("", pd.NA)
        matricula = df["MATRICULA"].replace("", pd.NA)
        df["VEHICULO_ID"] = chasis.where(chasis.notna(), matricula)

    vehicle_level_cols = [
        "FECHA",
        "CANAL",
        "DIAS EN TALLER",
        "COMPANIA",
        "NRO SINIESTRO",
        "CHASIS",
        "KILOMETRAJE",
        "MATRICULA",
        "NOMBRE CLIENTE",
        "TELEFONO",
        "MARCA",
        "MODELO",
        "STATUS DEL VEHICULO",
        "VENTA",
    ]
    grouped = df.groupby("VEHICULO_ID", dropna=False, sort=False)
    for col in vehicle_level_cols:
        if col in df.columns:
            df[col] = grouped[col].transform(lambda series: series.replace("", pd.NA).ffill().bfill())

    if "PROVEEDOR_NORMALIZADO" not in df.columns:
        df["PROVEEDOR_NORMALIZADO"] = df["PROVEEDOR"].apply(slug_text)

    if "PROVEEDOR_DISPLAY" not in df.columns:
        df["PROVEEDOR_DISPLAY"] = df["PROVEEDOR"].apply(provider_display_label)

    if "MAGNA_ADJUDICADO" not in df.columns:
        df["MAGNA_ADJUDICADO"] = df["PROVEEDOR_NORMALIZADO"] == "MAGNA"

    if "PIEZA_RESULTADO" not in df.columns:
        df["PIEZA_RESULTADO"] = df["PROVEEDOR"].apply(classify_piece_result)

    if "PIEZA_ENTREGADA" not in df.columns:
        df["PIEZA_ENTREGADA"] = df["STATUS DEL REPUESTO"].apply(slug_text) == "ENTREGADO"

    if "DIAS EFECTIVOS PIEZA" not in df.columns:
        fallback_days = (pd.Timestamp(date.today()) - df["FECHA"]).dt.days
        fallback_days = fallback_days.where(fallback_days >= 0)
        df["DIAS EFECTIVOS PIEZA"] = df["DIAS EN TALLER"].where(df["DIAS EN TALLER"].notna(), fallback_days)

    if "MOTIVO_NORMALIZADO" not in df.columns:
        df["MOTIVO_NORMALIZADO"] = df["MOTIVO"].apply(normalize_motivo_particular)

    if "GARANTIA_ESTADO_PIEZA" not in df.columns:
        df["GARANTIA_ESTADO_PIEZA"] = df.apply(
            lambda row: classify_warranty_status(row.get("CANAL"), row.get("KILOMETRAJE"), row.get("VENTA")),
            axis=1,
        )

    if "SEMAFORO_STOCK_PIEZA" not in df.columns:
        df["SEMAFORO_STOCK_PIEZA"] = df.apply(
            lambda row: classify_stock_wait(row.get("DIAS EFECTIVOS PIEZA"))
            if row.get("MOTIVO_NORMALIZADO") == "NO HAY STOCK"
            else "SIN DATO",
            axis=1,
        )

    if "SOURCE_SHEET" not in df.columns:
        df["SOURCE_SHEET"] = ""

    return df


def read_sheet_smart(file_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    raw_df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None)
    header_row = detect_header_row(raw_df)
    data = raw_df.iloc[header_row + 1 :].copy()
    data.columns = standardize_columns(raw_df.iloc[header_row].tolist())
    data = data.dropna(how="all").reset_index(drop=True)
    return clean_taller_dataframe(data, sheet_name)


@st.cache_data(show_spinner=False)
def load_workbook(file_bytes: bytes, cache_version: str = WORKBOOK_CACHE_VERSION) -> dict[str, pd.DataFrame]:
    _ = cache_version
    workbook = pd.ExcelFile(BytesIO(file_bytes))
    data: dict[str, pd.DataFrame] = {}
    for sheet_name in workbook.sheet_names:
        try:
            data[sheet_name] = read_sheet_smart(file_bytes, sheet_name)
        except Exception:
            continue
    return data


def build_vehicle_summary(df: pd.DataFrame) -> pd.DataFrame:
    df = ensure_analysis_columns(df)

    if df.empty:
        return pd.DataFrame(
            columns=[
                "VEHICULO_ID",
                "FECHA",
                "CANAL",
                "DIAS EN TALLER",
                "DIAS EFECTIVOS",
                "SEMAFORO TALLER",
                "COMPANIA",
                "NRO SINIESTRO",
                "PROVEEDOR",
                "CHASIS",
                "KILOMETRAJE",
                "VENTA",
                "ANTIGUEDAD ANOS",
                "GARANTIA ESTADO",
                "MATRICULA",
                "NOMBRE CLIENTE",
                "MARCA",
                "MODELO",
                "STATUS DEL VEHICULO",
                "PIEZAS SOLICITADAS",
                "PIEZAS GANADAS",
                "PIEZAS PERDIDAS",
                "PIEZAS PENDIENTES",
                "PIEZAS SIN PROVEEDOR",
                "PIEZAS SIN STOCK",
                "PIEZAS MUY CARAS",
                "DIAS ESPERA STOCK",
                "SEMAFORO STOCK",
                "PIEZAS ENTREGADAS",
                "PIEZAS SIN ENTREGAR",
                "ESPERANDO REPUESTOS",
                "MOTIVOS DETECTADOS",
                "LISTA REPUESTOS",
                "MAGNA ADJUDICADO",
            ]
        )

    grouped = df.groupby("VEHICULO_ID", dropna=False, sort=False)
    summary = grouped.agg(
        FECHA=("FECHA", "min"),
        CANAL=("CANAL", first_non_empty),
        DIAS_DECLARADOS=("DIAS EN TALLER", "max"),
        COMPANIA=("COMPANIA", first_non_empty),
        NRO_SINIESTRO=("NRO SINIESTRO", first_non_empty),
        PROVEEDOR=("PROVEEDOR", vehicle_provider_label),
        CHASIS=("CHASIS", first_non_empty),
        KILOMETRAJE=("KILOMETRAJE", "max"),
        VENTA=("VENTA", "min"),
        MATRICULA=("MATRICULA", first_non_empty),
        NOMBRE_CLIENTE=("NOMBRE CLIENTE", first_non_empty),
        MARCA=("MARCA", first_non_empty),
        MODELO=("MODELO", first_non_empty),
        STATUS_DEL_VEHICULO=("STATUS DEL VEHICULO", first_non_empty),
        PIEZAS_SOLICITADAS=("REPUESTOS SOLICITADO", count_non_empty),
        PIEZAS_GANADAS=("PIEZA_RESULTADO", lambda series: int(sum(value == "GANADA MAGNA" for value in series))),
        PIEZAS_PERDIDAS=("PIEZA_RESULTADO", lambda series: int(sum(value == "PERDIDA" for value in series))),
        PIEZAS_PENDIENTES=("PIEZA_RESULTADO", lambda series: int(sum(value == "SIN PROVEEDOR" for value in series))),
        PIEZAS_SIN_PROVEEDOR=("PROVEEDOR", lambda series: int(sum(not normalize_text(value) for value in series))),
        PIEZAS_SIN_STOCK=("MOTIVO_NORMALIZADO", lambda series: int(sum(value == "NO HAY STOCK" for value in series))),
        PIEZAS_MUY_CARAS=("MOTIVO_NORMALIZADO", lambda series: int(sum(value == "MUY CARO Y HAY STOCK" for value in series))),
        PIEZAS_ENTREGADAS=("PIEZA_ENTREGADA", "sum"),
        MOTIVOS_DETECTADOS=("MOTIVO_NORMALIZADO", unique_join),
        LISTA_REPUESTOS=("REPUESTOS SOLICITADO", unique_join),
        MAGNA_ADJUDICADO=("MAGNA_ADJUDICADO", "max"),
    ).reset_index()

    today = pd.Timestamp(date.today())
    fallback_days = (today - summary["FECHA"]).dt.days
    fallback_days = fallback_days.where(fallback_days >= 0)
    summary["DIAS EFECTIVOS"] = summary["DIAS_DECLARADOS"].where(summary["DIAS_DECLARADOS"].notna(), fallback_days)
    summary["DIAS EFECTIVOS"] = pd.to_numeric(summary["DIAS EFECTIVOS"], errors="coerce")
    summary["ANTIGUEDAD ANOS"] = ((pd.Timestamp(date.today()) - summary["VENTA"]).dt.days / 365.25).round(1)
    summary["GARANTIA ESTADO"] = summary.apply(
        lambda row: classify_warranty_status(row.get("CANAL"), row.get("KILOMETRAJE"), row.get("VENTA")),
        axis=1,
    )
    summary["PIEZAS SIN ENTREGAR"] = summary["PIEZAS_SOLICITADAS"] - summary["PIEZAS_ENTREGADAS"]
    status_slug = summary["STATUS_DEL_VEHICULO"].apply(slug_text)
    summary["ESPERANDO REPUESTOS"] = (
        status_slug.eq("ESPERANDO REPUESTOS")
        | (status_slug.eq("EN TALLER") & summary["PIEZAS SIN ENTREGAR"].gt(0))
    )
    summary["SEMAFORO TALLER"] = summary["DIAS EFECTIVOS"].apply(classify_semaforo)

    stock_wait_by_vehicle = (
        df[df["MOTIVO_NORMALIZADO"] == "NO HAY STOCK"]
        .groupby("VEHICULO_ID", dropna=False)["DIAS EFECTIVOS PIEZA"]
        .max()
    )
    summary["DIAS ESPERA STOCK"] = summary["VEHICULO_ID"].map(stock_wait_by_vehicle)
    summary["SEMAFORO STOCK"] = summary["DIAS ESPERA STOCK"].apply(classify_stock_wait)
    return summary


def provider_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["PROVEEDOR", "PIEZAS"])
    data = (
        df.groupby("PROVEEDOR_DISPLAY", dropna=False)
        .agg(PIEZAS=("REPUESTOS SOLICITADO", count_non_empty))
        .reset_index()
        .sort_values("PIEZAS", ascending=False)
    )
    return data.rename(columns={"PROVEEDOR_DISPLAY": "PROVEEDOR"})


def pieces_result_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["RESULTADO", "PIEZAS"])
    data = (
        df.groupby("PIEZA_RESULTADO", dropna=False)
        .agg(PIEZAS=("REPUESTOS SOLICITADO", count_non_empty))
        .reset_index()
        .rename(columns={"PIEZA_RESULTADO": "RESULTADO"})
    )
    data["RESULTADO"] = pd.Categorical(data["RESULTADO"], categories=PIECE_RESULT_ORDER, ordered=True)
    return data.sort_values("RESULTADO")


def status_summary(vehicle_summary: pd.DataFrame) -> pd.DataFrame:
    if vehicle_summary.empty:
        return pd.DataFrame(columns=["STATUS DEL VEHICULO", "VEHICULOS"])
    data = (
        vehicle_summary.groupby("STATUS_DEL_VEHICULO", dropna=False)
        .agg(VEHICULOS=("VEHICULO_ID", "count"))
        .reset_index()
        .sort_values("VEHICULOS", ascending=False)
    )
    data["STATUS DEL VEHICULO"] = data["STATUS_DEL_VEHICULO"].replace("", "SIN ESTADO")
    return data[["STATUS DEL VEHICULO", "VEHICULOS"]]


def semaforo_summary(vehicle_summary: pd.DataFrame) -> pd.DataFrame:
    if vehicle_summary.empty:
        return pd.DataFrame(columns=["SEMAFORO", "VEHICULOS"])
    data = (
        vehicle_summary.groupby("SEMAFORO TALLER", dropna=False)
        .agg(VEHICULOS=("VEHICULO_ID", "count"))
        .reset_index()
        .rename(columns={"SEMAFORO TALLER": "SEMAFORO"})
    )
    data["SEMAFORO"] = pd.Categorical(data["SEMAFORO"], categories=SEMAFORO_ORDER, ordered=True)
    return data.sort_values("SEMAFORO")


def warranty_status_summary(vehicle_summary: pd.DataFrame) -> pd.DataFrame:
    if vehicle_summary.empty:
        return pd.DataFrame(columns=["GARANTIA ESTADO", "VEHICULOS"])
    data = (
        vehicle_summary.groupby("GARANTIA ESTADO", dropna=False)
        .agg(VEHICULOS=("VEHICULO_ID", "count"))
        .reset_index()
    )
    data["GARANTIA ESTADO"] = pd.Categorical(data["GARANTIA ESTADO"], categories=WARRANTY_STATUS_ORDER, ordered=True)
    return data.sort_values("GARANTIA ESTADO")


def motivo_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["MOTIVO", "PIEZAS"])
    data = (
        df.groupby("MOTIVO_NORMALIZADO", dropna=False)
        .agg(PIEZAS=("REPUESTOS SOLICITADO", count_non_empty))
        .reset_index()
        .rename(columns={"MOTIVO_NORMALIZADO": "MOTIVO"})
    )
    ordered = MOTIVO_PARTICULAR_ORDER + sorted([m for m in data["MOTIVO"].tolist() if m not in MOTIVO_PARTICULAR_ORDER])
    data["MOTIVO"] = pd.Categorical(data["MOTIVO"], categories=ordered, ordered=True)
    return data.sort_values("MOTIVO")


def stock_wait_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["SEMAFORO STOCK", "PIEZAS"])
    stock_df = df[df["MOTIVO_NORMALIZADO"] == "NO HAY STOCK"].copy()
    if stock_df.empty:
        return pd.DataFrame(columns=["SEMAFORO STOCK", "PIEZAS"])
    data = (
        stock_df.groupby("SEMAFORO_STOCK_PIEZA", dropna=False)
        .agg(PIEZAS=("REPUESTOS SOLICITADO", count_non_empty))
        .reset_index()
        .rename(columns={"SEMAFORO_STOCK_PIEZA": "SEMAFORO STOCK"})
    )
    data["SEMAFORO STOCK"] = pd.Categorical(data["SEMAFORO STOCK"], categories=STOCK_WAIT_ORDER, ordered=True)
    return data.sort_values("SEMAFORO STOCK")


def top_no_stock_wait(df: pd.DataFrame, max_rows: int = 15) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=[
                "CLIENTE",
                "CANAL",
                "CHASIS",
                "MATRICULA",
                "MARCA",
                "MODELO",
                "PROVEEDOR",
                "REPUESTO",
                "DIAS ESPERA",
                "SEMAFORO STOCK",
                "STATUS DEL VEHICULO",
            ]
        )
    stock_df = df[df["MOTIVO_NORMALIZADO"] == "NO HAY STOCK"].copy()
    if stock_df.empty:
        return pd.DataFrame(
            columns=[
                "CLIENTE",
                "CANAL",
                "CHASIS",
                "MATRICULA",
                "MARCA",
                "MODELO",
                "PROVEEDOR",
                "REPUESTO",
                "DIAS ESPERA",
                "SEMAFORO STOCK",
                "STATUS DEL VEHICULO",
            ]
        )
    stock_df["CLIENTE"] = stock_df["NOMBRE CLIENTE"]
    stock_df["CANAL"] = stock_df["CANAL"]
    stock_df["REPUESTO"] = stock_df["REPUESTOS SOLICITADO"]
    stock_df["DIAS ESPERA"] = stock_df["DIAS EFECTIVOS PIEZA"]
    stock_df["SEMAFORO STOCK"] = stock_df["SEMAFORO_STOCK_PIEZA"]
    stock_df["PROVEEDOR"] = stock_df["PROVEEDOR_DISPLAY"]
    stock_df["STATUS DEL VEHICULO"] = stock_df["STATUS DEL VEHICULO"]
    return stock_df.sort_values(["DIAS ESPERA", "CLIENTE"], ascending=[False, True])[
        [
            "CLIENTE",
            "CANAL",
            "CHASIS",
            "MATRICULA",
            "MARCA",
            "MODELO",
            "PROVEEDOR",
            "REPUESTO",
            "DIAS ESPERA",
            "SEMAFORO STOCK",
            "STATUS DEL VEHICULO",
        ]
    ].head(max_rows)


def brand_or_model_summary(vehicle_summary: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    if vehicle_summary.empty:
        return pd.DataFrame(columns=["CATEGORIA", "VEHICULOS"]), "Marca"

    if vehicle_summary["MARCA"].replace("", pd.NA).nunique(dropna=True) > 1:
        data = (
            vehicle_summary.groupby("MARCA", dropna=False)
            .agg(VEHICULOS=("VEHICULO_ID", "count"))
            .reset_index()
            .sort_values("VEHICULOS", ascending=False)
        )
        data["CATEGORIA"] = data["MARCA"].replace("", "SIN MARCA")
        return data[["CATEGORIA", "VEHICULOS"]], "Marca"

    data = (
        vehicle_summary.groupby("MODELO", dropna=False)
        .agg(VEHICULOS=("VEHICULO_ID", "count"))
        .reset_index()
        .sort_values("VEHICULOS", ascending=False)
    )
    data["CATEGORIA"] = data["MODELO"].replace("", "SIN MODELO")
    return data[["CATEGORIA", "VEHICULOS"]], "Modelo"


def top_vehicles_by_delay(vehicle_summary: pd.DataFrame) -> pd.DataFrame:
    if vehicle_summary.empty:
        return pd.DataFrame(columns=["VEHICULO", "DIAS EFECTIVOS", "SEMAFORO TALLER"])
    data = vehicle_summary.copy()
    data["VEHICULO"] = data.apply(
        lambda row: f"{row['MATRICULA'] or row['CHASIS']} | {row['MODELO'] or 'Sin modelo'}",
        axis=1,
    )
    return data.sort_values(["DIAS EFECTIVOS", "PIEZAS SIN ENTREGAR"], ascending=[False, False])[
        ["VEHICULO", "DIAS EFECTIVOS", "SEMAFORO TALLER", "PIEZAS SIN ENTREGAR", "STATUS_DEL_VEHICULO"]
    ]


def format_date(value: object) -> str:
    if isinstance(value, pd.Timestamp) and not pd.isna(value):
        return value.strftime("%Y-%m-%d")
    return normalize_text(value)


def metric_card(title: str, value: object, help_text: str) -> None:
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{title}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-help">{help_text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_status_box(css_class: str, message: str) -> None:
    st.markdown(
        f'<div class="{css_class}">{message}</div>',
        unsafe_allow_html=True,
    )


def horizontal_bar(
    df: pd.DataFrame,
    category_col: str,
    value_col: str,
    title: str,
    color: str,
) -> None:
    if df.empty:
        st.info("No hay datos para este grafico con los filtros actuales.")
        return

    chart = (
        alt.Chart(df)
        .mark_bar(cornerRadiusTopRight=7, cornerRadiusBottomRight=7)
        .encode(
            x=alt.X(f"{value_col}:Q", title=""),
            y=alt.Y(f"{category_col}:N", sort="-x", title=""),
            color=alt.value(color),
            tooltip=[category_col, value_col],
        )
        .properties(height=max(220, 36 * len(df)), title=title)
    )
    st.altair_chart(chart, use_container_width=True)


def excel_table_name(base: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]", "", base.title())
    if not cleaned:
        cleaned = "Tabla"
    if cleaned[0].isdigit():
        cleaned = f"T{cleaned}"
    return cleaned[:25]


def autosize_worksheet(ws, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), max_width)


def style_data_sheet(ws, table_name: str, accent_color: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor=accent_color)
    header_font = Font(color="FFFFFF", bold=True)
    body_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = body_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = body_border

    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    autosize_worksheet(ws)


def write_dataframe_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, accent_color: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.book[sheet_name]
    style_data_sheet(ws, excel_table_name(sheet_name), accent_color)


def draw_kpi_box(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    value: object,
    subtitle: str,
    fill_color: str,
) -> None:
    end_col = start_col + 2
    border = Border(
        left=Side(style="thin", color="D6DEE8"),
        right=Side(style="thin", color="D6DEE8"),
        top=Side(style="thin", color="D6DEE8"),
        bottom=Side(style="thin", color="D6DEE8"),
    )
    fill = PatternFill("solid", fgColor=fill_color)

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=end_col)
    ws.merge_cells(start_row=start_row + 3, start_column=start_col, end_row=start_row + 3, end_column=end_col)

    for row in range(start_row, start_row + 4):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=start_row, column=start_col).value = title
    ws.cell(row=start_row, column=start_col).font = Font(color="FFFFFF", bold=True, size=11)
    ws.cell(row=start_row + 1, column=start_col).value = value
    ws.cell(row=start_row + 1, column=start_col).font = Font(color="FFFFFF", bold=True, size=18)
    ws.cell(row=start_row + 3, column=start_col).value = subtitle
    ws.cell(row=start_row + 3, column=start_col).font = Font(color="E2E8F0", size=9)


def write_meta_block(ws, start_row: int, start_col: int, title: str, rows: list[tuple[str, str]], accent_color: str) -> None:
    end_col = start_col + 4
    title_fill = PatternFill("solid", fgColor=accent_color)
    body_fill = PatternFill("solid", fgColor="F8FAFC")
    border = Border(
        left=Side(style="thin", color="D6DEE8"),
        right=Side(style="thin", color="D6DEE8"),
        top=Side(style="thin", color="D6DEE8"),
        bottom=Side(style="thin", color="D6DEE8"),
    )

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = title
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=11)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = border

    for offset, (label, value) in enumerate(rows, start=1):
        row = start_row + offset
        ws.cell(row=row, column=start_col).value = label
        ws.cell(row=row, column=start_col).font = Font(bold=True, color="0F172A")
        ws.cell(row=row, column=start_col).fill = body_fill
        ws.cell(row=row, column=start_col).border = border
        ws.cell(row=row, column=start_col).alignment = Alignment(vertical="top")

        ws.merge_cells(start_row=row, start_column=start_col + 1, end_row=row, end_column=end_col)
        value_cell = ws.cell(row=row, column=start_col + 1)
        value_cell.value = value
        value_cell.fill = body_fill
        value_cell.border = border
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_bar_chart(
    target_ws,
    source_ws,
    title: str,
    category_col: int,
    value_col: int,
    anchor: str,
    color: str,
    width: float = 10.5,
    height: float = 6.5,
) -> None:
    if source_ws.max_row < 2:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.height = height
    chart.width = width
    chart.varyColors = False
    chart.legend = None
    chart.gapWidth = 45

    data = Reference(source_ws, min_col=value_col, min_row=1, max_row=source_ws.max_row)
    categories = Reference(source_ws, min_col=category_col, min_row=2, max_row=source_ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    if chart.series:
        chart.series[0].graphicalProperties.solidFill = color
        chart.series[0].graphicalProperties.line.solidFill = color

    target_ws.add_chart(chart, anchor)


def build_executive_sheet(
    workbook,
    report_meta: dict[str, object],
    source_sheet_names: dict[str, str],
) -> None:
    ws = workbook.create_sheet("Reporte ejecutivo", 0)
    ws.sheet_view.showGridLines = False

    for column_letter, width in {
        "A": 16,
        "B": 16,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 16,
    }.items():
        ws.column_dimensions[column_letter].width = width

    ws.merge_cells("A1:L2")
    ws["A1"] = "Reporte Ejecutivo Taller Magna"
    ws["A1"].font = Font(size=22, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:L3")
    ws["A3"] = (
        f"Archivo: {report_meta['archivo']} | Hoja: {report_meta['hoja']} | "
        f"Generado: {report_meta['generado']}"
    )
    ws["A3"].font = Font(size=10, color="475569")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    for idx, kpi in enumerate(report_meta["kpis"][:4]):
        draw_kpi_box(ws, 5, 1 + (idx * 3), kpi["title"], kpi["value"], kpi["subtitle"], kpi["color"])

    for idx, kpi in enumerate(report_meta["kpis"][4:8]):
        draw_kpi_box(ws, 10, 1 + (idx * 3), kpi["title"], kpi["value"], kpi["subtitle"], kpi["color"])

    write_meta_block(ws, 15, 1, "Contexto del reporte", report_meta["context_rows"], "0F766E")
    write_meta_block(ws, 15, 7, "Filtros aplicados", report_meta["filter_rows"], "1D4ED8")

    for chart_cfg in report_meta["chart_configs"]:
        add_bar_chart(
            ws,
            workbook[source_sheet_names[chart_cfg["key"]]],
            chart_cfg["title"],
            1,
            2,
            chart_cfg["anchor"],
            chart_cfg["color"],
        )

    ws.merge_cells("A54:L55")
    ws["A54"] = (
        "Este archivo incluye un resumen ejecutivo con graficos y hojas de detalle para vehiculos, repuestos, "
        "proveedores, estados y demoras. Los graficos reflejan exactamente los filtros aplicados al momento de descargar."
    )
    ws["A54"].font = Font(size=10, color="475569", italic=True)
    ws["A54"].alignment = Alignment(wrap_text=True, vertical="top")


def dataframe_to_excel_bytes(
    vehicle_summary_df: pd.DataFrame,
    repuestos_df: pd.DataFrame,
    provider_df: pd.DataFrame,
    delay_df: pd.DataFrame,
    report_meta: dict[str, object],
    chart_tables: list[dict[str, object]],
) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        write_dataframe_sheet(writer, vehicle_summary_df, "Vehiculos", "0F766E")
        write_dataframe_sheet(writer, repuestos_df, "Repuestos", "0F172A")
        write_dataframe_sheet(writer, provider_df, "Proveedores", "2563EB")
        write_dataframe_sheet(writer, delay_df, "Demoras", "DC2626")

        chart_sheet_names: dict[str, str] = {}
        for chart_cfg in chart_tables:
            write_dataframe_sheet(writer, chart_cfg["df"], chart_cfg["sheet_name"], chart_cfg["accent_color"])
            chart_sheet_names[chart_cfg["key"]] = chart_cfg["sheet_name"]

        build_executive_sheet(
            writer.book,
            report_meta,
            chart_sheet_names,
        )
    return buffer.getvalue()


def get_input_file() -> tuple[bytes | None, str]:
    uploaded_file = st.sidebar.file_uploader("Subir Excel del taller", type=["xlsx", "xls"])
    if uploaded_file is not None:
        return uploaded_file.getvalue(), uploaded_file.name
    if DEFAULT_EXCEL_PATH.exists():
        return DEFAULT_EXCEL_PATH.read_bytes(), DEFAULT_EXCEL_PATH.name
    return None, ""


def build_search_mask(vehicle_summary: pd.DataFrame, search_text: str) -> pd.Series:
    if not search_text:
        return pd.Series(True, index=vehicle_summary.index)
    query = search_text.strip().lower()
    searchable_cols = ["CHASIS", "MATRICULA", "NOMBRE_CLIENTE", "MODELO", "LISTA_REPUESTOS", "NRO_SINIESTRO"]
    combined = (
        vehicle_summary[searchable_cols]
        .fillna("")
        .astype(str)
        .agg(" ".join, axis=1)
        .str.lower()
    )
    return combined.str.contains(query, na=False)


inject_css()

st.markdown(
    """
    <div class="title-card">
        <h1>Dashboard <span>Taller Magna</span></h1>
        <p>Resumen ejecutivo de siniestros, garantias, motivos del cliente y demoras por falta de repuestos.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

file_bytes, active_file_name = get_input_file()

if not file_bytes:
    st.error("No se encontro un archivo Excel para analizar.")
    st.stop()

workbook_data = load_workbook(file_bytes, WORKBOOK_CACHE_VERSION)
valid_sheets = [sheet for sheet in PREFERRED_SHEETS if sheet in workbook_data and not workbook_data[sheet].empty]
valid_sheets += [sheet for sheet in workbook_data if sheet not in valid_sheets and not workbook_data[sheet].empty]

if not valid_sheets:
    st.error("No se detectaron hojas con estructura valida de taller en el archivo.")
    st.stop()

default_sheet = "SINIESTROS" if "SINIESTROS" in valid_sheets else valid_sheets[0]

st.sidebar.subheader("Fuente")
selected_sheet = st.sidebar.selectbox(
    "Hoja a analizar",
    valid_sheets,
    index=valid_sheets.index(default_sheet),
    help="SINIESTROS sirve para siniestros y adjudicaciones. PARTICULAR Y GARANTIAS sirve para garantia, motivos y esperas sin stock.",
)

raw_df = ensure_analysis_columns(workbook_data[selected_sheet].copy())
vehicle_summary_df = build_vehicle_summary(raw_df)
is_particular_mode = slug_text(selected_sheet) == "PARTICULAR Y GARANTIAS"
semaforo_filter_column = "SEMAFORO STOCK" if is_particular_mode else "SEMAFORO TALLER"
semaforo_filter_label = "Semaforo de espera sin stock" if is_particular_mode else "Semaforo de demora"
semaforo_filter_order = STOCK_WAIT_ORDER if is_particular_mode else SEMAFORO_ORDER
waiting_filter_label = (
    "Solo vehiculos con piezas sin stock"
    if is_particular_mode
    else "Solo vehiculos esperando repuestos"
)

st.sidebar.markdown("---")
st.sidebar.subheader("Filtros")

brand_options = sorted([value for value in vehicle_summary_df["MARCA"].dropna().unique().tolist() if value])
status_options = sorted([value for value in vehicle_summary_df["STATUS_DEL_VEHICULO"].dropna().unique().tolist() if value])
provider_options = sorted([value for value in raw_df["PROVEEDOR_DISPLAY"].dropna().unique().tolist() if value])
semaforo_options = [
    value for value in semaforo_filter_order if value in vehicle_summary_df[semaforo_filter_column].astype(str).tolist()
]

selected_brands = st.sidebar.multiselect("Marca", brand_options, default=brand_options)
selected_status = st.sidebar.multiselect("Estado del vehiculo", status_options, default=status_options)
selected_providers = st.sidebar.multiselect("Proveedor", provider_options, default=provider_options)
selected_semaforo = st.sidebar.multiselect(semaforo_filter_label, semaforo_options, default=semaforo_options)
only_waiting_parts = st.sidebar.checkbox(waiting_filter_label, value=False)
only_magna = st.sidebar.checkbox("Solo proveedor MAGNA" if is_particular_mode else "Solo piezas ganadas por MAGNA", value=False)
search_text = st.sidebar.text_input("Buscar cliente, chasis, matricula o repuesto")

vehicle_mask = pd.Series(True, index=vehicle_summary_df.index)
if brand_options:
    vehicle_mask &= vehicle_summary_df["MARCA"].isin(selected_brands)
if status_options:
    vehicle_mask &= vehicle_summary_df["STATUS_DEL_VEHICULO"].isin(selected_status)
if selected_semaforo:
    vehicle_mask &= vehicle_summary_df[semaforo_filter_column].isin(selected_semaforo)
if only_waiting_parts:
    vehicle_mask &= (
        vehicle_summary_df["PIEZAS_SIN_STOCK"].gt(0)
        if is_particular_mode
        else vehicle_summary_df["ESPERANDO REPUESTOS"]
    )
vehicle_mask &= build_search_mask(vehicle_summary_df, search_text)

filtered_vehicle_summary = vehicle_summary_df.loc[vehicle_mask].copy()
filtered_vehicle_ids = filtered_vehicle_summary["VEHICULO_ID"].tolist()
filtered_df = raw_df[raw_df["VEHICULO_ID"].isin(filtered_vehicle_ids)].copy()

if provider_options:
    filtered_df = filtered_df[filtered_df["PROVEEDOR_DISPLAY"].isin(selected_providers)].copy()
if only_magna:
    filtered_df = filtered_df[filtered_df["MAGNA_ADJUDICADO"]].copy()

allowed_vehicle_ids = filtered_df["VEHICULO_ID"].dropna().unique().tolist()
filtered_vehicle_summary = filtered_vehicle_summary[filtered_vehicle_summary["VEHICULO_ID"].isin(allowed_vehicle_ids)].copy()

if is_particular_mode:
    st.info(
        "En PARTICULAR Y GARANTIAS el analisis prioriza garantia por kilometraje y fecha de venta, motivos del cliente y espera por falta de stock."
    )
    if (
        raw_df["KILOMETRAJE"].dropna().empty
        or raw_df["VENTA"].dropna().empty
        or raw_df["MOTIVO"].apply(normalize_text).eq("").all()
    ):
        st.warning(
            "Faltan datos en KILOMETRAJE, VENTA o MOTIVO. Cuando esas columnas no estan cargadas, la garantia queda en revision y el motivo se muestra como SIN MOTIVO."
        )
elif selected_sheet != "SINIESTROS":
    st.info(
        "Estas viendo una hoja distinta a SINIESTROS. Para revisar piezas ganadas y perdidas por MAGNA, la referencia mas util suele ser SINIESTROS."
    )

total_vehicles = len(filtered_vehicle_summary)
vehicles_in_shop = int(filtered_vehicle_summary["STATUS_DEL_VEHICULO"].apply(slug_text).eq("EN TALLER").sum())
vehicles_waiting_parts = int(filtered_vehicle_summary["ESPERANDO REPUESTOS"].sum())
avg_days = float(filtered_vehicle_summary["DIAS EFECTIVOS"].dropna().mean()) if total_vehicles else 0
critical_vehicles = int(filtered_vehicle_summary["SEMAFORO TALLER"].eq("CRITICA").sum())

total_pieces = int(count_non_empty(filtered_df["REPUESTOS SOLICITADO"]))
won_pieces = int((filtered_df["PIEZA_RESULTADO"] == "GANADA MAGNA").sum())
lost_pieces = int((filtered_df["PIEZA_RESULTADO"] == "PERDIDA").sum())
pending_pieces = int((filtered_df["PIEZA_RESULTADO"] == "SIN PROVEEDOR").sum())
delivered_pieces = int(filtered_df["PIEZA_ENTREGADA"].sum())

effectiveness = (won_pieces / (won_pieces + lost_pieces) * 100) if (won_pieces + lost_pieces) else 0

warranty_df = warranty_status_summary(filtered_vehicle_summary)
motivo_df = motivo_summary(filtered_df)
stock_wait_df = stock_wait_summary(filtered_df)
no_stock_delay_df = top_no_stock_wait(filtered_df)

garantia_vigente = int(filtered_vehicle_summary["GARANTIA ESTADO"].eq("GARANTIA VIGENTE").sum())
garantia_no_aplica = int(filtered_vehicle_summary["GARANTIA ESTADO"].astype(str).str.startswith("SIN GARANTIA").sum())
garantia_revisar = int(filtered_vehicle_summary["GARANTIA ESTADO"].astype(str).str.startswith("REVISAR").sum())
garantia_km_out = int(filtered_vehicle_summary["GARANTIA ESTADO"].astype(str).str.contains(">100.000 KM", regex=False).sum())
garantia_age_out = int(filtered_vehicle_summary["GARANTIA ESTADO"].astype(str).str.contains(">3 ANOS", regex=False).sum())
pieces_no_stock = int((filtered_df["MOTIVO_NORMALIZADO"] == "NO HAY STOCK").sum())
pieces_expensive_stock = int((filtered_df["MOTIVO_NORMALIZADO"] == "MUY CARO Y HAY STOCK").sum())
critical_stock_pieces = int((filtered_df["SEMAFORO_STOCK_PIEZA"] == ">55 DIAS - CRITICA").sum())
avg_stock_wait = (
    float(filtered_df.loc[filtered_df["MOTIVO_NORMALIZADO"] == "NO HAY STOCK", "DIAS EFECTIVOS PIEZA"].dropna().mean())
    if pieces_no_stock
    else 0
)

hero_message = (
    f"Archivo activo: <strong>{active_file_name}</strong> | Hoja: <strong>{selected_sheet}</strong> | "
    f"Vehiculos analizados: <strong>{total_vehicles}</strong> | Piezas analizadas: <strong>{total_pieces}</strong>"
)
st.markdown(
    f"""
    <div class="hero-card">
        <div class="hero-title">Resumen operativo del taller</div>
        <div class="hero-text">{hero_message}</div>
    </div>
    """,
    unsafe_allow_html=True,
)

if is_particular_mode and critical_stock_pieces > 0:
    render_status_box(
        "status-alert",
        f"Hay {critical_stock_pieces} piezas sin stock en situacion critica. Si superan 55 dias conviene cambiar de proveedor u otra solucion.",
    )
elif is_particular_mode and garantia_revisar > 0:
    render_status_box(
        "status-mid",
        f"Hay {garantia_revisar} vehiculos con garantia a revisar por kilometraje o fecha de venta.",
    )
elif (not is_particular_mode) and critical_vehicles > 0:
    render_status_box(
        "status-alert",
        f"Hay {critical_vehicles} vehiculos en situacion critica (71 dias o mas). Conviene revisarlos primero.",
    )
elif (is_particular_mode and pieces_no_stock > 0) or ((not is_particular_mode) and vehicles_waiting_parts > 0):
    render_status_box(
        "status-mid",
        (
            f"Hay {pieces_no_stock} piezas sin stock y conviene revisar la pestana de demoras para priorizar."
            if is_particular_mode
            else f"Hay {vehicles_waiting_parts} vehiculos esperando repuestos y conviene revisar la pestana de demoras para priorizar."
        ),
    )
else:
    render_status_box("status-good", "No se detectan alertas relevantes con los filtros actuales.")

piece_result_df = pieces_result_summary(filtered_df)
provider_df = provider_summary(filtered_df)
status_df = status_summary(filtered_vehicle_summary)
semaforo_df = semaforo_summary(filtered_vehicle_summary)
brand_df, brand_label = brand_or_model_summary(filtered_vehicle_summary)
delay_top_df = top_vehicles_by_delay(filtered_vehicle_summary).head(15)

tabs = st.tabs(
    [
        "Resumen garantia" if is_particular_mode else "Resumen ejecutivo",
        "Garantias" if is_particular_mode else "Piezas ganadas / perdidas",
        "Stock y motivos" if is_particular_mode else "Demoras en taller",
        "Detalle vehiculos",
        "Detalle repuestos",
    ]
)

with tabs[0]:
    metric_cols_top = st.columns(4)
    with metric_cols_top[0]:
        metric_card("Vehiculos analizados", total_vehicles, "Vehiculos incluidos despues de aplicar filtros.")
    with metric_cols_top[1]:
        metric_card(
            "Garantias vigentes" if is_particular_mode else "Vehiculos en taller",
            garantia_vigente if is_particular_mode else vehicles_in_shop,
            "Vehiculos garantia dentro de 100 mil km y hasta 3 anos desde la venta."
            if is_particular_mode
            else "Vehiculos cuyo estado actual figura como EN TALLER.",
        )
    with metric_cols_top[2]:
        metric_card(
            "Sin garantia" if is_particular_mode else "Esperando repuestos",
            garantia_no_aplica if is_particular_mode else vehicles_waiting_parts,
            "Vehiculos que exceden kilometraje o antiguedad permitida."
            if is_particular_mode
            else "Vehiculos frenados por repuestos sin entregar.",
        )
    with metric_cols_top[3]:
        metric_card(
            "Revisar datos" if is_particular_mode else "Dias promedio en taller",
            garantia_revisar if is_particular_mode else f"{avg_days:.0f}",
            "Falta kilometraje o fecha de venta para confirmar garantia."
            if is_particular_mode
            else "Usa DIAS EN TALLER o lo estima desde FECHA si esta vacio.",
        )

    metric_cols_bottom = st.columns(4)
    with metric_cols_bottom[0]:
        metric_card(
            "Piezas sin stock" if is_particular_mode else "Piezas ganadas MAGNA",
            pieces_no_stock if is_particular_mode else won_pieces,
            "Piezas donde el motivo indica NO HAY STOCK."
            if is_particular_mode
            else "Piezas con proveedor adjudicado igual a MAGNA.",
        )
    with metric_cols_bottom[1]:
        metric_card(
            "Muy caro y hay stock" if is_particular_mode else "Piezas perdidas",
            pieces_expensive_stock if is_particular_mode else lost_pieces,
            "Cliente no cambia la pieza por precio teniendo stock."
            if is_particular_mode
            else "Piezas adjudicadas a otros proveedores.",
        )
    with metric_cols_bottom[2]:
        metric_card(
            "Espera critica" if is_particular_mode else "Piezas sin proveedor",
            critical_stock_pieces if is_particular_mode else pending_pieces,
            "Piezas sin stock con espera superior a 55 dias."
            if is_particular_mode
            else "Filas donde la columna Proveedor esta vacia.",
        )
    with metric_cols_bottom[3]:
        metric_card(
            "Espera promedio" if is_particular_mode else "% efectividad",
            f"{avg_stock_wait:.0f}" if is_particular_mode else f"{effectiveness:.1f}%",
            "Promedio de dias de espera en piezas sin stock."
            if is_particular_mode
            else "Ganadas sobre ganadas + perdidas.",
        )

    chart_col_1, chart_col_2 = st.columns(2)
    with chart_col_1:
        horizontal_bar(
            warranty_df if is_particular_mode else piece_result_df,
            "GARANTIA ESTADO" if is_particular_mode else "RESULTADO",
            "VEHICULOS" if is_particular_mode else "PIEZAS",
            "Estado de garantia" if is_particular_mode else "Resultado de piezas",
            "#0f766e",
        )
    with chart_col_2:
        horizontal_bar(
            motivo_df if is_particular_mode else status_df,
            "MOTIVO" if is_particular_mode else "STATUS DEL VEHICULO",
            "PIEZAS" if is_particular_mode else "VEHICULOS",
            "Motivos del cliente" if is_particular_mode else "Vehiculos por estado",
            "#1d4ed8",
        )

    chart_col_3, chart_col_4 = st.columns(2)
    with chart_col_3:
        horizontal_bar(
            stock_wait_df if is_particular_mode else semaforo_df,
            "SEMAFORO STOCK" if is_particular_mode else "SEMAFORO",
            "PIEZAS" if is_particular_mode else "VEHICULOS",
            "Espera por no stock" if is_particular_mode else "Semaforo de demora",
            "#b45309",
        )
    with chart_col_4:
        horizontal_bar(brand_df, "CATEGORIA", "VEHICULOS", f"Vehiculos por {brand_label.lower()}", "#0f172a")

with tabs[1]:
    st.markdown(
        '<div class="tab-note">En esta pestana se resume la garantia por kilometraje y fecha de venta.</div>'
        if is_particular_mode
        else '<div class="tab-note">En esta pestana se ve el resultado comercial por pieza: ganada por MAGNA, perdida o sin proveedor asignado.</div>',
        unsafe_allow_html=True,
    )

    metric_cols = st.columns(4)
    with metric_cols[0]:
        metric_card(
            "Garantia vigente" if is_particular_mode else "Piezas analizadas",
            garantia_vigente if is_particular_mode else total_pieces,
            "Vehiculos aptos para garantia." if is_particular_mode else "Total de repuestos dentro de los filtros actuales.",
        )
    with metric_cols[1]:
        metric_card(
            "Fuera por km" if is_particular_mode else "Ganadas MAGNA",
            garantia_km_out if is_particular_mode else won_pieces,
            "Vehiculos que superan 100.000 km." if is_particular_mode else "Piezas adjudicadas a MAGNA.",
        )
    with metric_cols[2]:
        metric_card(
            "Fuera por venta" if is_particular_mode else "Perdidas",
            garantia_age_out if is_particular_mode else lost_pieces,
            "Vehiculos con venta mayor a 3 anos." if is_particular_mode else "Piezas adjudicadas a otros proveedores.",
        )
    with metric_cols[3]:
        metric_card(
            "Revisar datos" if is_particular_mode else "Sin proveedor",
            garantia_revisar if is_particular_mode else pending_pieces,
            "Faltan datos para decidir garantia." if is_particular_mode else "Piezas con la columna Proveedor vacia.",
        )

    chart_col_1, chart_col_2 = st.columns(2)
    with chart_col_1:
        horizontal_bar(
            warranty_df if is_particular_mode else piece_result_df,
            "GARANTIA ESTADO" if is_particular_mode else "RESULTADO",
            "VEHICULOS" if is_particular_mode else "PIEZAS",
            "Resumen de garantia" if is_particular_mode else "Piezas ganadas, perdidas y sin proveedor",
            "#0f766e",
        )
    with chart_col_2:
        horizontal_bar(
            status_df if is_particular_mode else provider_df.head(10),
            "STATUS DEL VEHICULO" if is_particular_mode else "PROVEEDOR",
            "VEHICULOS" if is_particular_mode else "PIEZAS",
            "Vehiculos por estado" if is_particular_mode else "Top proveedores por piezas",
            "#2563eb",
        )

with tabs[2]:
    st.markdown(
        '<div class="tab-note">Para piezas sin stock: 0 a 40 dias dentro del plazo, 41 a 55 llamada de atencion, mas de 55 dias situacion critica.</div>'
        if is_particular_mode
        else '<div class="tab-note">La demora se clasifica asi: 0 a 30 normal, 31 a 45 atencion, 46 a 70 demora alta, 71 o mas critica.</div>',
        unsafe_allow_html=True,
    )

    metric_cols = st.columns(4)
    with metric_cols[0]:
        metric_card(
            "Sin stock" if is_particular_mode else "Normal",
            pieces_no_stock if is_particular_mode else int(filtered_vehicle_summary["SEMAFORO TALLER"].eq("NORMAL").sum()),
            "Piezas cuya espera depende de reposicion." if is_particular_mode else "0 a 30 dias.",
        )
    with metric_cols[1]:
        metric_card(
            "Muy caro y hay stock" if is_particular_mode else "Atencion",
            pieces_expensive_stock if is_particular_mode else int(filtered_vehicle_summary["SEMAFORO TALLER"].eq("ATENCION").sum()),
            "Cliente no cambia la pieza por precio." if is_particular_mode else "31 a 45 dias.",
        )
    with metric_cols[2]:
        metric_card(
            "Critica >55 dias" if is_particular_mode else "Demora alta",
            critical_stock_pieces if is_particular_mode else int(filtered_vehicle_summary["SEMAFORO TALLER"].eq("DEMORA ALTA").sum()),
            "Conviene cambiar de proveedor u otra solucion." if is_particular_mode else "46 a 70 dias.",
        )
    with metric_cols[3]:
        metric_card(
            "Dias promedio espera" if is_particular_mode else "Critica",
            f"{avg_stock_wait:.0f}" if is_particular_mode else critical_vehicles,
            "Promedio solo en piezas sin stock." if is_particular_mode else "71 dias o mas.",
        )

    chart_col_1, chart_col_2 = st.columns(2)
    with chart_col_1:
        horizontal_bar(
            motivo_df if is_particular_mode else semaforo_df,
            "MOTIVO" if is_particular_mode else "SEMAFORO",
            "PIEZAS" if is_particular_mode else "VEHICULOS",
            "Motivos detectados" if is_particular_mode else "Vehiculos por semaforo",
            "#dc2626",
        )
    with chart_col_2:
        if is_particular_mode:
            horizontal_bar(stock_wait_df, "SEMAFORO STOCK", "PIEZAS", "Semaforo de espera por no stock", "#ea580c")
        else:
            waiting_df = filtered_vehicle_summary[filtered_vehicle_summary["ESPERANDO REPUESTOS"]].copy()
            waiting_semaforo_df = semaforo_summary(waiting_df)
            horizontal_bar(waiting_semaforo_df, "SEMAFORO", "VEHICULOS", "Esperando repuestos por semaforo", "#ea580c")

    st.subheader("Piezas con mayor espera por falta de stock" if is_particular_mode else "Vehiculos mas demorados")
    if is_particular_mode:
        st.dataframe(no_stock_delay_df, use_container_width=True, hide_index=True)
    else:
        delay_display = delay_top_df.rename(
            columns={
                "DIAS EFECTIVOS": "DIAS EN TALLER",
                "SEMAFORO TALLER": "SEMAFORO",
                "PIEZAS SIN ENTREGAR": "PIEZAS SIN ENTREGAR",
                "STATUS_DEL_VEHICULO": "STATUS DEL VEHICULO",
            }
        )
        st.dataframe(delay_display, use_container_width=True, hide_index=True)

vehicle_display = filtered_vehicle_summary.copy()
vehicle_display["FECHA"] = vehicle_display["FECHA"].apply(format_date)
vehicle_display["VENTA"] = vehicle_display["VENTA"].apply(format_date)
vehicle_display["MAGNA ADJUDICADO"] = vehicle_display["MAGNA_ADJUDICADO"].map({True: "SI", False: "NO"})
vehicle_display["ESPERANDO REPUESTOS"] = vehicle_display["ESPERANDO REPUESTOS"].map({True: "SI", False: "NO"})
vehicle_display = vehicle_display.rename(
    columns={
        "CANAL": "CANAL",
        "DIAS_DECLARADOS": "DIAS EN TALLER CARGADOS",
        "DIAS EFECTIVOS": "DIAS EN TALLER",
        "NRO_SINIESTRO": "NRO SINIESTRO",
        "NOMBRE_CLIENTE": "NOMBRE CLIENTE",
        "STATUS_DEL_VEHICULO": "STATUS DEL VEHICULO",
        "PIEZAS_SOLICITADAS": "PIEZAS SOLICITADAS",
        "PIEZAS_GANADAS": "PIEZAS GANADAS",
        "PIEZAS_PERDIDAS": "PIEZAS PERDIDAS",
        "PIEZAS_PENDIENTES": "PIEZAS PENDIENTES",
        "PIEZAS_SIN_PROVEEDOR": "PIEZAS SIN PROVEEDOR",
        "PIEZAS_SIN_STOCK": "PIEZAS SIN STOCK",
        "PIEZAS_MUY_CARAS": "PIEZAS MUY CARAS",
        "PIEZAS_ENTREGADAS": "PIEZAS ENTREGADAS",
        "PIEZAS SIN ENTREGAR": "PIEZAS SIN ENTREGAR",
        "SEMAFORO TALLER": "SEMAFORO",
        "SEMAFORO STOCK": "SEMAFORO STOCK",
        "GARANTIA ESTADO": "GARANTIA ESTADO",
        "MOTIVOS_DETECTADOS": "MOTIVOS DETECTADOS",
        "LISTA_REPUESTOS": "LISTA REPUESTOS",
    }
)
summary_columns = [
    "FECHA",
    "COMPANIA",
    "NRO SINIESTRO",
    "PROVEEDOR",
    "MAGNA ADJUDICADO",
    "CHASIS",
    "MATRICULA",
    "NOMBRE CLIENTE",
    "MARCA",
    "MODELO",
    "STATUS DEL VEHICULO",
    "ESPERANDO REPUESTOS",
    "SEMAFORO",
    "DIAS EN TALLER CARGADOS",
    "DIAS EN TALLER",
    "PIEZAS SOLICITADAS",
    "PIEZAS GANADAS",
    "PIEZAS PERDIDAS",
    "PIEZAS PENDIENTES",
    "PIEZAS SIN PROVEEDOR",
    "PIEZAS ENTREGADAS",
    "PIEZAS SIN ENTREGAR",
    "LISTA REPUESTOS",
]
if is_particular_mode:
    summary_columns = [
        "FECHA",
        "CANAL",
        "PROVEEDOR",
        "CHASIS",
        "KILOMETRAJE",
        "VENTA",
        "ANTIGUEDAD ANOS",
        "GARANTIA ESTADO",
        "MATRICULA",
        "NOMBRE CLIENTE",
        "MARCA",
        "MODELO",
        "STATUS DEL VEHICULO",
        "SEMAFORO STOCK",
        "DIAS ESPERA STOCK",
        "DIAS EN TALLER CARGADOS",
        "DIAS EN TALLER",
        "PIEZAS SOLICITADAS",
        "PIEZAS SIN STOCK",
        "PIEZAS MUY CARAS",
        "PIEZAS SIN PROVEEDOR",
        "MOTIVOS DETECTADOS",
        "LISTA REPUESTOS",
    ]
summary_export = vehicle_display[summary_columns].copy()

repuestos_display = filtered_df.copy()
repuestos_display["FECHA"] = repuestos_display["FECHA"].apply(format_date)
repuestos_display["FECHA ENTREGA PIEZA"] = repuestos_display["FECHA ENTREGA PIEZA"].apply(format_date)
repuestos_display["VENTA"] = repuestos_display["VENTA"].apply(format_date)
repuestos_display["MAGNA ADJUDICADO"] = repuestos_display["MAGNA_ADJUDICADO"].map({True: "SI", False: "NO"})
repuestos_display["PROVEEDOR"] = repuestos_display["PROVEEDOR_DISPLAY"]
repuestos_display = repuestos_display.rename(
    columns={
        "PIEZA_RESULTADO": "RESULTADO PIEZA",
        "GARANTIA_ESTADO_PIEZA": "GARANTIA ESTADO PIEZA",
        "DIAS EFECTIVOS PIEZA": "DIAS ESPERA PIEZA",
        "SEMAFORO_STOCK_PIEZA": "SEMAFORO STOCK",
        "MOTIVO_NORMALIZADO": "MOTIVO NORMALIZADO",
    }
)
repuestos_columns = [
    "FECHA",
    "COMPANIA",
    "NRO SINIESTRO",
    "PROVEEDOR",
    "MAGNA ADJUDICADO",
    "RESULTADO PIEZA",
    "CHASIS",
    "MATRICULA",
    "NOMBRE CLIENTE",
    "MARCA",
    "MODELO",
    "CODIGO",
    "REPUESTOS SOLICITADO",
    "STATUS DEL REPUESTO",
    "STATUS DEL VEHICULO",
    "FECHA ENTREGA PIEZA",
    "COMENTARIOS",
]
if is_particular_mode:
    repuestos_columns = [
        "FECHA",
        "CANAL",
        "PROVEEDOR",
        "CHASIS",
        "KILOMETRAJE",
        "VENTA",
        "GARANTIA ESTADO PIEZA",
        "MATRICULA",
        "NOMBRE CLIENTE",
        "MARCA",
        "MODELO",
        "CODIGO",
        "REPUESTOS SOLICITADO",
        "MONTO PIEZA",
        "MONTO M OBRA",
        "MOTIVO",
        "MOTIVO NORMALIZADO",
        "DIAS ESPERA PIEZA",
        "SEMAFORO STOCK",
        "STATUS DEL REPUESTO",
        "STATUS DEL VEHICULO",
        "FECHA ENTREGA PIEZA",
        "COMENTARIOS",
    ]
repuestos_export = repuestos_display[repuestos_columns].copy()

if is_particular_mode:
    delay_export = no_stock_delay_df.copy()
    chart_tables = [
        {"key": "chart_1", "df": warranty_df, "sheet_name": "Garantia vehiculos", "accent_color": "0F766E"},
        {"key": "chart_2", "df": motivo_df, "sheet_name": "Motivos cliente", "accent_color": "1D4ED8"},
        {"key": "chart_3", "df": stock_wait_df, "sheet_name": "Espera sin stock", "accent_color": "B45309"},
        {"key": "chart_4", "df": brand_df, "sheet_name": "Marca modelo", "accent_color": "0F172A"},
    ]
    report_meta = {
        "archivo": active_file_name,
        "hoja": selected_sheet,
        "generado": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
        "context_rows": [
            ("Vehiculos analizados", str(total_vehicles)),
            ("Garantias vigentes", str(garantia_vigente)),
            ("Sin garantia", str(garantia_no_aplica)),
            ("Revisar datos", str(garantia_revisar)),
            ("Piezas sin stock", str(pieces_no_stock)),
            ("Muy caro y hay stock", str(pieces_expensive_stock)),
        ],
        "filter_rows": [
            ("Marca", ", ".join(selected_brands) if selected_brands else "Todas"),
            ("Estado vehiculo", ", ".join(selected_status) if selected_status else "Todos"),
            ("Proveedor", ", ".join(selected_providers) if selected_providers else "Todos"),
            ("Semaforo stock", ", ".join(selected_semaforo) if selected_semaforo else "Todos"),
            ("Solo piezas sin stock", "SI" if only_waiting_parts else "NO"),
            ("Solo proveedor MAGNA", "SI" if only_magna else "NO"),
            ("Busqueda", search_text if search_text else "Sin texto"),
        ],
        "kpis": [
            {"title": "Vehiculos analizados", "value": total_vehicles, "subtitle": "Vehiculos incluidos despues de aplicar filtros.", "color": "0F766E"},
            {"title": "Garantias vigentes", "value": garantia_vigente, "subtitle": "Dentro de 100 mil km y hasta 3 años.", "color": "1D4ED8"},
            {"title": "Sin garantia", "value": garantia_no_aplica, "subtitle": "Fuera por kilometraje o antigüedad.", "color": "B45309"},
            {"title": "Revisar datos", "value": garantia_revisar, "subtitle": "Falta kilometraje o fecha de venta.", "color": "0F172A"},
            {"title": "Piezas sin stock", "value": pieces_no_stock, "subtitle": "Motivo NO HAY STOCK.", "color": "0F766E"},
            {"title": "Muy caro y hay stock", "value": pieces_expensive_stock, "subtitle": "Cliente no cambia por precio.", "color": "1D4ED8"},
            {"title": "Espera critica", "value": critical_stock_pieces, "subtitle": "No stock por encima de 55 dias.", "color": "B45309"},
            {"title": "Espera promedio", "value": f"{avg_stock_wait:.0f}", "subtitle": "Promedio de espera sin stock.", "color": "0F172A"},
        ],
        "chart_configs": [
            {"key": "chart_1", "title": "Estado de garantia", "anchor": "A24", "color": "0F766E"},
            {"key": "chart_2", "title": "Motivos del cliente", "anchor": "G24", "color": "1D4ED8"},
            {"key": "chart_3", "title": "Espera por no stock", "anchor": "A39", "color": "B45309"},
            {"key": "chart_4", "title": f"Vehiculos por {brand_label.lower()}", "anchor": "G39", "color": "0F172A"},
        ],
    }
else:
    delay_export = delay_top_df.rename(
        columns={
            "DIAS EFECTIVOS": "DIAS EN TALLER",
            "SEMAFORO TALLER": "SEMAFORO",
            "PIEZAS SIN ENTREGAR": "PIEZAS SIN ENTREGAR",
            "STATUS_DEL_VEHICULO": "STATUS DEL VEHICULO",
        }
    ).copy()
    chart_tables = [
        {"key": "chart_1", "df": piece_result_df, "sheet_name": "Resultado piezas", "accent_color": "0F766E"},
        {"key": "chart_2", "df": status_df, "sheet_name": "Estados vehiculo", "accent_color": "1D4ED8"},
        {"key": "chart_3", "df": semaforo_df, "sheet_name": "Semaforo taller", "accent_color": "B45309"},
        {"key": "chart_4", "df": brand_df, "sheet_name": "Marca modelo", "accent_color": "0F172A"},
    ]
    report_meta = {
        "archivo": active_file_name,
        "hoja": selected_sheet,
        "generado": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
        "context_rows": [
            ("Vehiculos analizados", str(total_vehicles)),
            ("Piezas analizadas", str(total_pieces)),
            ("Piezas ganadas MAGNA", str(won_pieces)),
            ("Piezas perdidas", str(lost_pieces)),
            ("Piezas sin proveedor", str(pending_pieces)),
            ("Efectividad", f"{effectiveness:.1f}%"),
        ],
        "filter_rows": [
            ("Marca", ", ".join(selected_brands) if selected_brands else "Todas"),
            ("Estado vehiculo", ", ".join(selected_status) if selected_status else "Todos"),
            ("Proveedor", ", ".join(selected_providers) if selected_providers else "Todos"),
            ("Semaforo taller", ", ".join(selected_semaforo) if selected_semaforo else "Todos"),
            ("Solo esperando repuestos", "SI" if only_waiting_parts else "NO"),
            ("Solo piezas ganadas MAGNA", "SI" if only_magna else "NO"),
            ("Busqueda", search_text if search_text else "Sin texto"),
        ],
        "kpis": [
            {"title": "Vehiculos analizados", "value": total_vehicles, "subtitle": "Vehiculos incluidos despues de aplicar filtros.", "color": "0F766E"},
            {"title": "Vehiculos en taller", "value": vehicles_in_shop, "subtitle": "Estado actual EN TALLER.", "color": "1D4ED8"},
            {"title": "Esperando repuestos", "value": vehicles_waiting_parts, "subtitle": "Vehiculos frenados por piezas sin entregar.", "color": "B45309"},
            {"title": "Dias promedio", "value": f"{avg_days:.0f}", "subtitle": "Promedio usando dias cargados o estimados.", "color": "0F172A"},
            {"title": "Piezas ganadas MAGNA", "value": won_pieces, "subtitle": "Proveedor adjudicado igual a MAGNA.", "color": "0F766E"},
            {"title": "Piezas perdidas", "value": lost_pieces, "subtitle": "Adjudicadas a otros proveedores.", "color": "1D4ED8"},
            {"title": "Piezas sin proveedor", "value": pending_pieces, "subtitle": "Filas donde la columna Proveedor esta vacia.", "color": "B45309"},
            {"title": "Efectividad", "value": f"{effectiveness:.1f}%", "subtitle": "Ganadas sobre ganadas + perdidas.", "color": "0F172A"},
        ],
        "chart_configs": [
            {"key": "chart_1", "title": "Resultado de piezas", "anchor": "A24", "color": "0F766E"},
            {"key": "chart_2", "title": "Vehiculos por estado", "anchor": "G24", "color": "1D4ED8"},
            {"key": "chart_3", "title": "Semaforo de demora", "anchor": "A39", "color": "B45309"},
            {"key": "chart_4", "title": f"Vehiculos por {brand_label.lower()}", "anchor": "G39", "color": "0F172A"},
        ],
    }

download_bytes = dataframe_to_excel_bytes(
    summary_export,
    repuestos_export,
    provider_df,
    delay_export,
    report_meta,
    chart_tables,
)

with tabs[3]:
    st.subheader("Detalle por vehiculo")
    vehicle_sort_cols = ["SEMAFORO", "DIAS EN TALLER"] if "SEMAFORO" in summary_export.columns else ["GARANTIA ESTADO", "DIAS EN TALLER"]
    vehicle_sort_asc = [True, False]
    st.dataframe(
        summary_export.sort_values(vehicle_sort_cols, ascending=vehicle_sort_asc),
        use_container_width=True,
        hide_index=True,
    )

with tabs[4]:
    st.subheader("Detalle de repuestos")
    st.dataframe(
        repuestos_export.sort_values(["PROVEEDOR", "CHASIS", "REPUESTOS SOLICITADO"], ascending=[True, True, True]),
        use_container_width=True,
        hide_index=True,
    )

st.download_button(
    "Descargar reporte ejecutivo en Excel",
    data=download_bytes,
    file_name=f"resumen_taller_magna_{selected_sheet.lower().replace(' ', '_')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
